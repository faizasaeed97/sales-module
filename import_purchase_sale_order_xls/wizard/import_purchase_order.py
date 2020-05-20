# -*- coding: utf-8 -*-

import os
import csv
import base64
import tempfile
from odoo.exceptions import UserError
from odoo import api, fields, models, _, SUPERUSER_ID
from datetime import datetime, timedelta, date
from xlrd import open_workbook
import openpyxl
import io
import xlrd, mmap, xlwt


class ImportPurchaseOrder(models.TransientModel):
    _name = "wizard.import.purchase.order"

    file_data = fields.Binary('Archive', required=True, )
    file_name = fields.Char('File Name')

    def import_button(self):
        if not self.csv_validator(self.file_name):
            raise UserError(_("The file must be an .xls/.xlsx extension"))
        file_path = tempfile.gettempdir() + '/file.xlsx'
        data = self.file_data
        # f = open(file_path,'wb')
        # # f.write(data.decode('base64'))
        # s2 = base64.b64encode(data);
        # s3=base64.b64decode(s2)
        # f.write(s3)
        #
        #
        # f.close()

        decoded_data = base64.b64decode(self.file_data)
        xls_filelike = io.BytesIO(decoded_data)
        workbook = openpyxl.load_workbook(xls_filelike)

        # workbook = xlrd.open_workbook(file_path, on_demand = True)
        worksheet = workbook.active
        first_row = []  # The row where we stock the name of the column
        for col in range(worksheet.max_column):
            first_row.append(worksheet.cell(1, col + 1).value)
        # transform the workbook to a list of dictionaries
        archive_lines = []
        for row in range(1, worksheet.max_row):
            elm = {}
            for col in range(worksheet.max_column):
                elm[first_row[col]] = worksheet.cell(row + 1, col + 1).value

            archive_lines.append(elm)

        productx = self.env['product.template']
        # contract = self.env['hr.product']
        categ = self.env['product.category']
        uom_obj = self.env['uom.uom']
        p_brand=self.env['brand']
        p_type=self.env['product.type']
        p_color=self.env['product.color']
        p_length=self.env['product.length']
        p_width=self.env['product.width']
        p_height=self.env['product.height']
        p_made = self.env['product.made']
        p_model= self.env['product.made']
        stock_quant = self.env['stock.quant']
        # p_height = self.env['product.height']

        # self.valid_columns_keys(archive_lines)
        # self.valid_product_code(archive_lines, product_obj)
        # self.valid_prices(archive_lines)

        # vals = {
        #     'partner_id': self.partner_id.id,
        #     'date_planned': datetime.now(),
        # }
        # purchase_order_id = purchase_order_obj.create(vals)
        cont = 0
        for line in archive_lines:
            cont += 1
            # sub = str(line.get('Sub-Category', ""))
            cat = str(line.get('Product Category', ""))
            um = str(line.get('UOM -3', ""))
            Brand =p_brand.search([('name','=',str(line.get(u'Brand',"")))])
            if not Brand:
                Brand=p_brand.create({'name':str(line.get(u'Brand',""))})

            Made = p_made.search([('name','=',str(line.get(u'Made', "")))])
            if not Made:
                Made=p_made.create({'name':str(line.get(u'Made', ""))})

            color =p_color.search([('name','=',str(line.get(u'color', "")))])
            if not color:
                color=p_color.create({'name':str(line.get(u'color', ""))})

            Length =p_length.search([('name','=',line.get(u'Length', 0))])
            if not Length:
                Length=p_length.create({'name':line.get(u'Length', 0)})

            Width =p_width.search([('name','=', line.get(u'Width', 0))])
            if not Width:
                Width=p_width.create({'name':line.get(u'Width', 0)})

            Height =p_height.search([('name','=', line.get(u'Height', 0))])
            if not Height:
                Height=p_height.create({'name':line.get(u'Height', 0)})

            Type =p_type.search([('name','=',str(line.get(u'Product Type', "")))])
            if not Type:
                Type=p_type.create({'name':str(line.get(u'Product Type', ""))})
            cp = line.get(u'cp', 0.0)
            sp=line.get(u'sp', 0.0)

            description = str(line.get(u'description', ""))
            Material = str(line.get(u'Material', ""))
            Qty= line.get(u'Qty - 3', 0)
            Model = str(line.get(u'Model', ""))




            uom = uom_obj.search([('name', '=', um)])
            cat_id = categ.search([('name', '=', cat)], limit=1)



            # taxes = product_id.supplier_taxes_id.filtered(lambda r: not product_id.company_id or r.company_id == product_id.company_id)
            # tax_ids = taxes.ids
            if line.get('description', False):

                # if not cat_id:
                #    cat_id = categ.search([('name', '=', sub), ('parent_id.name', '=', cat)])

                if not cat_id:
                    cat_id = categ.create({
                        'name': cat,
                    })

                # if not cat_id:
                #     cat_id=1

                # if line.get('Travel Allowance', 0.0) == ' ':
                #     ha = 0.0
                # else:
                #     ha = (line.get('Housing Allowance', 0.0)) or 0.0
                #
                # if line.get('Travel Allowance', 0.0) == ' ':
                #     ta = 0.0
                # else:
                #     ta = line.get('Travel Allowance', 0.0) or 0.0
                #
                # if line.get('Basic Salary', 0.0) == ' ':
                #     wa = 0.0
                # else:
                #     wa = line.get('Basic Salary', 0.0) or 0.0
                #
                # if line.get('GOSI Salary Deduction', 0.0) == ' ':
                #     gos = 0.0
                # else:
                #     gos = line.get('GOSI Salary Deduction', 0.0) or 0.0
                name=cat_id.name  +' '

                if Length:
                    if Length.name:
                        name +=Length.name + ' '
                    else:
                        name += ""
                if Width:
                    if Width.name:
                       name += 'X' + Width.name + ' '
                    else:
                        name += ""
                if Height:
                    if Height.name:
                       name += 'X' +Height.name + ' '
                    else:
                        name += ""

                if color and color.name !=None or False:
                    name += color.name + ' '
                else:
                    name += ""

                if Type and Type.name != None:
                    name+=Type.name + ' '
                else:
                    name += ""
                if Made and Made.name != None:
                    name+=Made.name + ' '
                else:
                    name += ""
                if Brand and Brand.name !=None:
                    name+=Brand.name
                else:
                    name+=""




                vals = {
                    'name': name,
                    'uom_id': 34,
                    'uom_po_id': 34,
                    'categ_id': cat_id.id,
                    'type': 'product',
                    'raw_mat': True,
                    'model':Model,
                    'Material':Material,
                    'description':description,
                    'sequence':cont,

                }


                has = productx.search([('name', '=', name)])
                if len(has) > 0:
                    has.write(vals)
                else:
                    try:
                      ct = self.env['product.template'].sudo().create(vals)
                      product=self.env['product.product'].search([('product_tmpl_id','=',ct.id)])
                      if product:
                         loc=self.env.ref('stock.stock_location_stock').id
                         stok= self.env['stock.quant'].sudo().create({
                              'product_id': product.id,
                              'location_id': loc,
                              'quantity': Qty,
                          })
                      else:
                          product=self.env['product.product'].create({'product_tmpl_id':ct.id})
                          if product:
                              loc = self.env.ref('stock.stock_location_stock').id
                              stok = self.env['stock.quant'].sudo().create({
                                  'product_id': product.id,
                                  'location_id': loc,
                                  'quantity': Qty,
                              })

                    except:
                        print("THIS IS THE SHIT------->",vals)

        # if self._context.get('open_order', False):
        #     return purchase_order_id.action_view_order(purchase_order_id.id)
        # return {'type': 'ir.actions.act_window_close'}

    #
    # def import_buttonxx(self):
    #     if not self.csv_validator(self.file_name):
    #         raise UserError(_("The file must be an .xls/.xlsx extension"))
    #     file_path = tempfile.gettempdir()+'/file.xlsx'
    #     data = self.file_data
    #     # f = open(file_path,'wb')
    #     # # f.write(data.decode('base64'))
    #     # s2 = base64.b64encode(data);
    #     # s3=base64.b64decode(s2)
    #     # f.write(s3)
    #     #
    #     #
    #     # f.close()
    #
    #     decoded_data = base64.b64decode(self.file_data)
    #     xls_filelike = io.BytesIO(decoded_data)
    #     workbook = openpyxl.load_workbook(xls_filelike)
    #
    #     # workbook = xlrd.open_workbook(file_path, on_demand = True)
    #     worksheet = workbook.active
    #     first_row = [] # The row where we stock the name of the column
    #     for col in range(worksheet.max_column):
    #         first_row.append(worksheet.cell(1,col+1).value)
    #     # transform the workbook to a list of dictionaries
    #     archive_lines = []
    #     for row in range(1, worksheet.max_row):
    #         elm = {}
    #         for col in range(worksheet.max_column):
    #             elm[first_row[col]]=worksheet.cell(row+1,col+1).value
    #
    #         archive_lines.append(elm)
    #
    #
    #
    #     employee = self.env['hr.employee']
    #     # contract = self.env['hr.product']
    #     # product_template_obj = self.env['product.template']
    #     hr_contract = self.env['hr.contract']
    #
    #
    #     # self.valid_columns_keys(archive_lines)
    #     # self.valid_product_code(archive_lines, product_obj)
    #     # self.valid_prices(archive_lines)
    #
    #     # vals = {
    #     #     'partner_id': self.partner_id.id,
    #     #     'date_planned': datetime.now(),
    #     # }
    #     # purchase_order_id = purchase_order_obj.create(vals)
    #     cont = 0
    #     for line in archive_lines:
    #         cont += 1
    #         code = str(line.get('Roll',""))
    #         emp = str(line.get('Name', ""))
    #         employe_id = employee.search([('identification_id','=',code),('name','=',emp)])
    #         # quantity = line.get(u'quantity',0)
    #         # price_unit = self.get_valid_price(line.get('price',""),cont)
    #         # product_uom = product_template_obj.search([('default_code','=',code)])
    #         # taxes = product_id.supplier_taxes_id.filtered(lambda r: not product_id.company_id or r.company_id == product_id.company_id)
    #         # tax_ids = taxes.ids
    #         if  employe_id:
    #             if line.get('Travel Allowance',0.0) ==' ':
    #                 ha=0.0
    #             else:
    #                 ha = (line.get('Housing Allowance', 0.0)) or 0.0
    #
    #             if line.get('Travel Allowance',0.0) ==' ':
    #                 ta=0.0
    #             else:
    #                 ta= line.get('Travel Allowance',0.0) or 0.0
    #
    #             if line.get('Basic Salary',0.0) == ' ':
    #                 wa=0.0
    #             else:
    #                 wa = line.get('Basic Salary', 0.0) or 0.0
    #
    #             if line.get('GOSI Salary Deduction',0.0) == ' ':
    #                 gos=0.0
    #             else:
    #                gos=line.get('GOSI Salary Deduction',0.0) or 0.0
    #
    #             vals = {
    #                 'date_start':  datetime.strptime(line.get('Date of Join',""), '%d/%m/%Y').date(),
    #                 'name': employe_id.name,
    #                 'employee_id':employe_id.id,
    #                 'department_id': employe_id.department_id.id,
    #                 'company_id':employe_id.company_id.id,
    #                 'sponsorship': line.get('Sponsor',""),
    #                 'housing_allowance':ha,
    #                 'travel_allowance': ta,
    #                 'wage': wa,
    #                 'gosi_Salary_Deduction':gos,
    #                 'state':'open'
    #             }
    #             print(vals)
    #             has=hr_contract.search([('name','=',employe_id.name),('employee_id','=',employe_id.id)])
    #             if len(has)>0:
    #                 has.write(vals)
    #             else:
    #                ct= self.env['hr.contract'].sudo().create(vals)
    #
    #     # if self._context.get('open_order', False):
    #     #     return purchase_order_id.action_view_order(purchase_order_id.id)
    #     # return {'type': 'ir.actions.act_window_close'}

    # @api.model
    # def valid_prices(self, archive_lines):
    #     cont = 0
    #     for line in archive_lines:
    #         cont += 1
    #         price = line.get('price',"")
    #         if price != "":
    #             price = str(price).replace("$","").replace(",",".")
    #         try:
    #             price_float = float(price)
    #         except:
    #             raise UserError("The price of the line item %s does not have an appropriate format, for example: '100.00' - '100'"%cont)
    #
    #     return True
    #
    # @api.model
    # def get_valid_price(self, price, cont):
    #     if price != "":
    #         price = str(price).replace("$","").replace(",",".")
    #     try:
    #         price_float = float(price)
    #         return price_float
    #     except:
    #         raise UserError("The price of the line item %s does not have an appropriate format, for example: '100.00' - '100'"%cont)
    #     return False
    #
    # @api.model
    # def valid_product_code(self, archive_lines, product_obj):
    #     cont=0
    #     for line in archive_lines:
    #         cont += 1
    #         code = str(line.get('code',"")).strip()
    #         product_id = product_obj.search([('default_code','=',code)])
    #         if len(product_id)>1:
    #             raise UserError("The product code of line %s is duplicated in the system."%cont)
    #         if not product_id:
    #             raise UserError("The product code of line %s can't be found in the system."%cont)
    #
    # @api.model
    # def valid_columns_keys(self, archive_lines):
    #     columns = archive_lines[0].keys()
    #     print("columns>>",columns)
    #     text = "The file must contain the following columns: code, quantity, and price. \n The following columns are not in the file:"; text2 = text
    #     if not 'code' in columns:
    #         text +="\n[ code ]"
    #     if not u'quantity' in columns:
    #         text +="\n[ quantity ]"
    #     if not 'price' in columns:
    #         text +="\n[ price ]"
    #     if text !=text2:
    #         raise UserError(text)
    #     return True

    @api.model
    def csv_validator(self, xml_name):
        name, extension = os.path.splitext(xml_name)
        return True if extension == '.xls' or extension == '.xlsx' else False




